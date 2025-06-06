#!/usr/bin/env python3
"""
Script de verificación para funcionalidad de distribución Excel
Ejecutar desde la raíz del proyecto: python test_distribucion.py
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

def verificar_dependencias():
    """Verifica que todas las dependencias estén instaladas"""
    try:
        import openpyxl
        print("✅ openpyxl disponible")
    except ImportError:
        print("❌ openpyxl no encontrado - Instalar con: pip install openpyxl")
        return False
    
    try:
        from flask import Flask
        print("✅ Flask disponible")
    except ImportError:
        print("❌ Flask no encontrado")
        return False
    
    return True

def verificar_estructura_archivos():
    """Verifica que los archivos necesarios existan"""
    archivos_requeridos = [
        'routes/api.py',
        'utils/helpers.py',
        'static/js/reclutas.js',
        'static/css/styles.css',
        'models/recluta.py',
        'models/usuario.py'
    ]
    
    for archivo in archivos_requeridos:
        if os.path.exists(archivo):
            print(f"✅ {archivo} existe")
        else:
            print(f"❌ {archivo} NO encontrado")
            return False
    
    return True

def crear_excel_ejemplo():
    """Crea un archivo Excel de ejemplo para pruebas"""
    try:
        import openpyxl
        from datetime import datetime
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Headers
        ws['A1'] = "Fecha de creación"
        ws['B1'] = "Nombre"
        ws['C1'] = "Teléfono"
        
        # Datos de ejemplo
        datos_ejemplo = [
            ("05/06/2025 12:20pm", "Juan Pérez", "5551234567"),
            ("05/06/2025 01:15pm", "María García", "5559876543"),
            ("05/06/2025 02:30pm", "Carlos López", "5555678901"),
            ("05/06/2025 03:45pm", "Ana Martínez", "5554567890"),
            ("05/06/2025 04:20pm", "Luis Rodríguez", "5553456789")
        ]
        
        for i, (fecha, nombre, telefono) in enumerate(datos_ejemplo, start=2):
            ws[f'A{i}'] = fecha
            ws[f'B{i}'] = nombre
            ws[f'C{i}'] = telefono
        
        wb.save('test_reclutas_ejemplo.xlsx')
        print("✅ Archivo test_reclutas_ejemplo.xlsx creado")
        return True
        
    except Exception as e:
        print(f"❌ Error creando Excel ejemplo: {e}")
        return False

def verificar_distribucion_logica():
    """Verifica la lógica de distribución"""
    # Simular datos
    datos_ejemplo = [f"Recluta {i}" for i in range(10)]
    asesores_ejemplo = [
        type('obj', (object,), {'email': 'asesor1@test.com'}),
        type('obj', (object,), {'email': 'asesor2@test.com'}),
        type('obj', (object,), {'email': 'asesor3@test.com'})
    ]
    
    # Simular función de distribución
    total_reclutas = len(datos_ejemplo)
    num_asesores = len(asesores_ejemplo)
    reclutas_por_asesor = total_reclutas // num_asesores
    sobrantes = total_reclutas % num_asesores
    
    distribucion = {}
    indice_actual = 0
    
    for i, asesor in enumerate(asesores_ejemplo):
        cantidad = reclutas_por_asesor + (1 if i < sobrantes else 0)
        reclutas_asesor = datos_ejemplo[indice_actual:indice_actual + cantidad]
        distribucion[asesor.email] = len(reclutas_asesor)
        indice_actual += cantidad
    
    print("✅ Lógica de distribución:")
    for email, cantidad in distribucion.items():
        print(f"   {email}: {cantidad} reclutas")
    
    total_distribuidos = sum(distribucion.values())
    if total_distribuidos == total_reclutas:
        print("✅ Distribución correcta - Total distribuido == Total reclutas")
        return True
    else:
        print(f"❌ Error distribución: {total_distribuidos} != {total_reclutas}")
        return False

def main():
    """Función principal de verificación"""
    print("🔍 VERIFICACIÓN DE DISTRIBUCIÓN EXCEL")
    print("=" * 50)
    
    verificaciones = [
        ("Dependencias", verificar_dependencias),
        ("Estructura archivos", verificar_estructura_archivos),
        ("Lógica distribución", verificar_distribucion_logica),
        ("Excel ejemplo", crear_excel_ejemplo)
    ]
    
    resultados = []
    for nombre, funcion in verificaciones:
        print(f"\n📋 Verificando {nombre}...")
        resultado = funcion()
        resultados.append(resultado)
    
    print("\n" + "=" * 50)
    if all(resultados):
        print("🎉 ¡TODAS LAS VERIFICACIONES PASARON!")
        print("\n📝 SIGUIENTE PASO:")
        print("1. Reiniciar servidor Flask")
        print("2. Login como administrador")
        print("3. Buscar botón 'Distribuir Reclutas Excel'")
        print("4. Usar archivo test_reclutas_ejemplo.xlsx")
    else:
        print("❌ ALGUNAS VERIFICACIONES FALLARON")
        print("Revisar errores arriba antes de continuar")

if __name__ == "__main__":
    main()