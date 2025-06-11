import os
import uuid
import magic
import hashlib
import logging
from PIL import Image
from werkzeug.utils import secure_filename
from flask import current_app, request
from datetime import datetime, date, timedelta
import json
import re
from collections import defaultdict

# 🛡️ CONFIGURACIÓN DE SEGURIDAD PARA ARCHIVOS
ALLOWED_EXTENSIONS = {
    'image': {'jpg', 'jpeg', 'png', 'gif', 'webp'},
    'document': {'pdf', 'doc', 'docx', 'txt'},
    'excel': {'xlsx', 'xls', 'csv'}
}

ALLOWED_MIME_TYPES = {
    'image': {
        'image/jpeg', 'image/png', 'image/gif', 'image/webp'
    },
    'document': {
        'application/pdf', 
        'application/msword',
        'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        'text/plain'
    },
    'excel': {
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/vnd.ms-excel',
        'text/csv'
    }
}

# Magic bytes para verificación de contenido real
MAGIC_BYTES = {
    'image/jpeg': [b'\xff\xd8\xff'],
    'image/png': [b'\x89PNG\r\n\x1a\n'],
    'image/gif': [b'GIF87a', b'GIF89a'],
    'image/webp': [b'RIFF', b'WEBP'],
    'application/pdf': [b'%PDF-'],
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': [b'PK\x03\x04'],
    'application/vnd.ms-excel': [b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'],
    'text/csv': [b''],  # CSV puede empezar con cualquier caracter
}

# Límites de tamaño por tipo (en bytes)
MAX_FILE_SIZES = {
    'image': 5 * 1024 * 1024,      # 5MB para imágenes
    'document': 10 * 1024 * 1024,   # 10MB para documentos
    'excel': 15 * 1024 * 1024      # 15MB para archivos Excel
}

# Rate limiting básico por IP
upload_attempts = defaultdict(list)
MAX_UPLOADS_PER_HOUR = 50


def is_rate_limited(ip_address):
    """
    Verifica si la IP ha excedido el límite de uploads por hora.
    
    Args:
        ip_address: Dirección IP del cliente
        
    Returns:
        bool: True si está limitado, False si puede continuar
    """
    now = datetime.utcnow()
    hour_ago = now - timedelta(hours=1)
    
    # Limpiar intentos antiguos
    upload_attempts[ip_address] = [
        timestamp for timestamp in upload_attempts[ip_address] 
        if timestamp > hour_ago
    ]
    
    # Verificar límite
    if len(upload_attempts[ip_address]) >= MAX_UPLOADS_PER_HOUR:
        current_app.logger.warning(f"Rate limit excedido para IP: {ip_address}")
        return True
    
    # Registrar nuevo intento
    upload_attempts[ip_address].append(now)
    return False


def validate_file_extension(filename, file_type):
    """
    Valida que la extensión del archivo esté permitida.
    
    Args:
        filename: Nombre del archivo
        file_type: Tipo de archivo esperado
        
    Returns:
        bool: True si es válida, False si no
    """
    if not filename or '.' not in filename:
        return False
    
    extension = filename.rsplit('.', 1)[1].lower()
    return extension in ALLOWED_EXTENSIONS.get(file_type, set())


def validate_mime_type(file_content, expected_type):
    """
    Valida el tipo MIME real del archivo usando python-magic.
    
    Args:
        file_content: Contenido del archivo en bytes
        expected_type: Tipo esperado ('image', 'document', 'excel')
        
    Returns:
        tuple: (is_valid, detected_mime_type)
    """
    try:
        # Detectar tipo MIME real
        detected_mime = magic.from_buffer(file_content, mime=True)
        
        # Verificar si está en los tipos permitidos
        allowed_mimes = ALLOWED_MIME_TYPES.get(expected_type, set())
        is_valid = detected_mime in allowed_mimes
        
        current_app.logger.info(f"MIME detectado: {detected_mime}, Esperado: {expected_type}, Válido: {is_valid}")
        
        return is_valid, detected_mime
        
    except Exception as e:
        current_app.logger.error(f"Error detectando tipo MIME: {str(e)}")
        return False, None


def validate_magic_bytes(file_content, mime_type):
    """
    Valida los magic bytes del archivo para prevenir bypass por extensión.
    
    Args:
        file_content: Contenido del archivo en bytes
        mime_type: Tipo MIME detectado
        
    Returns:
        bool: True si los magic bytes son válidos
    """
    if not file_content:
        return False
    
    magic_signatures = MAGIC_BYTES.get(mime_type, [])
    
    # CSV puede tener cualquier inicio, omitir validación
    if mime_type == 'text/csv':
        return True
    
    if not magic_signatures:
        current_app.logger.warning(f"No hay magic bytes definidos para {mime_type}")
        return True  # Permitir si no hay definición específica
    
    # Verificar si algún magic byte coincide
    for magic_signature in magic_signatures:
        if file_content.startswith(magic_signature):
            return True
    
    current_app.logger.warning(f"Magic bytes no coinciden para {mime_type}")
    return False


def sanitize_image(file_path, max_width=1920, max_height=1080):
    """
    Sanitiza una imagen redimensionándola y removiendo metadatos EXIF.
    
    Args:
        file_path: Ruta del archivo de imagen
        max_width: Ancho máximo permitido
        max_height: Alto máximo permitido
        
    Returns:
        bool: True si se sanitizó correctamente
    """
    try:
        with Image.open(file_path) as img:
            # Remover metadatos EXIF
            data = list(img.getdata())
            image_without_exif = Image.new(img.mode, img.size)
            image_without_exif.putdata(data)
            
            # Redimensionar si es necesario
            if img.width > max_width or img.height > max_height:
                img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
                current_app.logger.info(f"Imagen redimensionada a {img.size}")
            
            # Guardar imagen sanitizada
            if img.format == 'JPEG':
                image_without_exif.save(file_path, 'JPEG', quality=85, optimize=True)
            else:
                image_without_exif.save(file_path, img.format, optimize=True)
            
            return True
            
    except Exception as e:
        current_app.logger.error(f"Error sanitizando imagen: {str(e)}")
        return False


def generate_secure_filename(original_filename):
    """
    Genera un nombre de archivo seguro y único.
    
    Args:
        original_filename: Nombre original del archivo
        
    Returns:
        str: Nombre seguro generado
    """
    if not original_filename:
        return f"file_{uuid.uuid4().hex}"
    
    # Securizar nombre original
    secure_name = secure_filename(original_filename)
    
    # Remover caracteres problemáticos adicionales
    secure_name = re.sub(r'[^\w\-_\.]', '', secure_name)
    
    # Separar nombre y extensión
    name_parts = secure_name.rsplit('.', 1)
    if len(name_parts) == 2:
        name, extension = name_parts
        # Limitar longitud del nombre
        name = name[:50] if len(name) > 50 else name
        # Agregar timestamp para unicidad
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        unique_id = uuid.uuid4().hex[:8]
        return f"{name}_{timestamp}_{unique_id}.{extension.lower()}"
    else:
        # Sin extensión, agregar .bin por seguridad
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        unique_id = uuid.uuid4().hex[:8]
        return f"{secure_name}_{timestamp}_{unique_id}.bin"


def eliminar_archivo(ruta_relativa):
    """
    Elimina un archivo del servidor de forma segura.
    
    Args:
        ruta_relativa: Ruta relativa del archivo a eliminar
        
    Returns:
        dict: {'success': bool, 'message': str}
    """
    if not ruta_relativa:
        return {'success': False, 'message': 'No se proporcionó ruta de archivo.'}
    
    try:
        # Construir ruta completa
        ruta_completa = os.path.join(current_app.root_path, ruta_relativa)
        
        # Verificar que la ruta esté dentro del directorio de uploads (seguridad)
        upload_folder = os.path.join(current_app.root_path, current_app.config['UPLOAD_FOLDER'])
        if not os.path.abspath(ruta_completa).startswith(os.path.abspath(upload_folder)):
            current_app.logger.warning(f"Intento de eliminar archivo fuera de uploads: {ruta_relativa}")
            return {'success': False, 'message': 'Ruta de archivo no válida.'}
        
        # Eliminar archivo si existe
        if os.path.exists(ruta_completa):
            os.remove(ruta_completa)
            current_app.logger.info(f"Archivo eliminado: {ruta_relativa}")
            return {'success': True, 'message': 'Archivo eliminado correctamente.'}
        else:
            return {'success': False, 'message': 'Archivo no encontrado.'}
            
    except Exception as e:
        current_app.logger.error(f"Error al eliminar archivo {ruta_relativa}: {str(e)}")
        return {'success': False, 'message': f'Error al eliminar archivo: {str(e)}'}


class JSONEncoder(json.JSONEncoder):
    """
    Codificador JSON personalizado para manejar tipos de datos específicos
    como datetime y date.
    """
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        return super().default(obj)


def format_date(date_string, output_format='%d/%m/%Y'):
    """
    Formatea una fecha en string al formato especificado.
    
    Args:
        date_string: Fecha como string
        output_format: Formato de salida deseado
        
    Returns:
        Fecha formateada o None si hay error
    """
    if not date_string:
        return None
    
    try:
        # Intentar varios formatos de entrada comunes
        input_formats = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S']
        
        parsed_date = None
        for fmt in input_formats:
            try:
                parsed_date = datetime.strptime(date_string, fmt)
                break
            except ValueError:
                continue
        
        if parsed_date:
            return parsed_date.strftime(output_format)
        else:
            return date_string  # Devolver original si no se puede parsear
            
    except Exception as e:
        current_app.logger.warning(f"Error formateando fecha '{date_string}': {str(e)}")
        return date_string


def clean_uploaded_files(days_old=30):
    """
    Función de mantenimiento para limpiar archivos antiguos.
    
    Args:
        days_old: Días de antigüedad para considerar archivo como obsoleto
        
    Returns:
        dict: Estadísticas de limpieza
    """
    try:
        upload_folder = os.path.join(current_app.root_path, current_app.config['UPLOAD_FOLDER'])
        cutoff_date = datetime.utcnow() - timedelta(days=days_old)
        
        files_deleted = 0
        space_freed = 0
        
        for root, dirs, files in os.walk(upload_folder):
            for file in files:
                file_path = os.path.join(root, file)
                
                # Verificar fecha de modificación
                if os.path.getmtime(file_path) < cutoff_date.timestamp():
                    try:
                        file_size = os.path.getsize(file_path)
                        os.remove(file_path)
                        files_deleted += 1
                        space_freed += file_size
                        current_app.logger.info(f"Archivo antiguo eliminado: {file_path}")
                    except Exception as e:
                        current_app.logger.error(f"Error eliminando {file_path}: {str(e)}")
        
        return {
            'files_deleted': files_deleted,
            'space_freed_mb': round(space_freed / 1024 / 1024, 2),
            'success': True
        }
        
    except Exception as e:
        current_app.logger.error(f"Error en limpieza de archivos: {str(e)}")
        return {'success': False, 'error': str(e)}

def guardar_archivo(archivo, tipo):
    """
    Guarda un archivo en el servidor y devuelve la ruta relativa.
    tipo puede ser 'recluta' o 'usuario'
    
    Args:
        archivo: Objeto de archivo de Flask (request.files)
        tipo: Tipo de archivo ('recluta' o 'usuario')
        
    Returns:
        Ruta relativa del archivo guardado, o None si hay un error
    """
    if not archivo:
        return None
        
    try:
        # Generar nombre único para el archivo
        filename = secure_filename(archivo.filename)
        nombre_base, extension = os.path.splitext(filename)
        nombre_unico = f"{nombre_base}_{uuid.uuid4().hex}{extension}"
        
        # Crear subdirectorio si no existe
        directorio = os.path.join(current_app.config['UPLOAD_FOLDER'], tipo)
        if not os.path.exists(directorio):
            os.makedirs(directorio)
        
        # Guardar el archivo
        ruta_completa = os.path.join(directorio, nombre_unico)
        archivo.save(ruta_completa)
        
        # Devolver ruta relativa para guardar en BD
        return os.path.join(f"static/uploads/{tipo}", nombre_unico)
    except Exception as e:
        current_app.logger.error(f"Error al guardar archivo: {str(e)}")
        return None

def paginate_data(data, page, per_page):
    """
    Implementa paginación manual para una lista de datos.
    
    Args:
        data: Lista de datos a paginar
        page: Número de página (comienza en 1)
        per_page: Elementos por página
        
    Returns:
        Diccionario con datos paginados y metadatos
    """
    total = len(data)
    pages = (total + per_page - 1) // per_page  # Redondeo hacia arriba
    
    # Validar página
    page = max(1, min(page, pages if pages > 0 else 1))
    
    # Calcular índices
    start_idx = (page - 1) * per_page
    end_idx = min(start_idx + per_page, total)
    
    # Extraer elementos de la página actual
    items = data[start_idx:end_idx]
    
    return {
        'items': items,
        'page': page,
        'per_page': per_page,
        'total': total,
        'pages': pages,
        'has_prev': page > 1,
        'has_next': page < pages
    }

import openpyxl
from datetime import datetime
from models.recluta import Recluta
from models import db

def procesar_y_distribuir_excel(archivo, asesores):
    """
    ✅ VERSIÓN CORREGIDA: Procesa archivo Excel y distribuye reclutas automáticamente entre asesores.
    
    FIXES APLICADOS:
    - Mapeo correcto de columnas respetando celdas vacías
    - Búsqueda directa en fila completa sin filtrado
    - Validación robusta de headers flexibles
    - Manejo mejorado de errores con detalles específicos
    
    Args:
        archivo: Archivo Excel uploadado
        asesores: Lista de usuarios asesores activos
        
    Returns:
        Dict con resultado detallado del procesamiento
    """
    try:
        # Leer Excel
        workbook = openpyxl.load_workbook(archivo, data_only=True)
        sheet = workbook.active
        
        # Obtener fila completa de headers (SIN FILTRAR)
        primera_fila = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        # ✅ BÚSQUEDA DIRECTA EN FILA COMPLETA (respeta columnas vacías)
        headers_requeridos = ["Fecha de creación", "Nombre", "Teléfono"]
        indices_encontrados = {}
        headers_disponibles = []
        
        # Mapear cada header requerido a su posición real en el Excel
        for i, cell in enumerate(primera_fila):
            cell_value = str(cell).strip() if cell is not None else ""
            if cell_value:  # Solo agregar headers no vacíos para logging
                headers_disponibles.append(cell_value)
            
            # Buscar headers requeridos por su posición exacta
            for header_req in headers_requeridos:
                if cell_value == header_req:
                    indices_encontrados[header_req] = i
                    break
        
        # Validar que se encontraron todos los headers requeridos
        headers_faltantes = [h for h in headers_requeridos if h not in indices_encontrados]
        if headers_faltantes:
            return {
                "success": False,
                "message": f"Headers faltantes: {headers_faltantes}. Disponibles: {headers_disponibles}",
                "headers_disponibles": headers_disponibles,
                "headers_requeridos": headers_requeridos
            }
        
        # ✅ USAR ÍNDICES REALES PARA ACCESO A DATOS
        indice_fecha = indices_encontrados["Fecha de creación"]
        indice_nombre = indices_encontrados["Nombre"]
        indice_telefono = indices_encontrados["Teléfono"]
        
        print(f"🔍 DEBUG - Índices mapeados: Fecha={indice_fecha}, Nombre={indice_nombre}, Teléfono={indice_telefono}")
        
        # Extraer y validar datos
        datos_excel = []
        errores = []
        telefonos_existentes = set()
        
        # Obtener teléfonos ya existentes en BD para evitar duplicados
        telefonos_bd = set(r[0] for r in db.session.query(Recluta.telefono).all())
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # ✅ ACCESO SEGURO CON ÍNDICES CORRECTOS
            fecha_str = row[indice_fecha] if len(row) > indice_fecha and row[indice_fecha] is not None else None
            nombre = row[indice_nombre] if len(row) > indice_nombre and row[indice_nombre] is not None else None
            telefono = row[indice_telefono] if len(row) > indice_telefono and row[indice_telefono] is not None else None
            
            # Limpiar y validar datos
            nombre = str(nombre).strip() if nombre is not None else ""
            telefono = str(telefono).strip() if telefono is not None else ""
            
            # Validaciones mejoradas
            error_fila = None
            
            if not nombre:
                error_fila = "Nombre vacío o inválido"
            elif not telefono:
                error_fila = "Teléfono vacío o inválido"
            elif telefono in telefonos_bd:
                error_fila = f"Teléfono {telefono} ya existe en BD"
            elif telefono in telefonos_existentes:
                error_fila = f"Teléfono {telefono} duplicado en Excel"
            
            if error_fila:
                errores.append({
                    "fila": row_num, 
                    "error": error_fila,
                    "datos": {"nombre": nombre, "telefono": telefono}
                })
            else:
                # Agregar a dataset válido
                telefonos_existentes.add(telefono)
                
                # Procesar fecha con múltiples formatos
                fecha_procesada = None
                if fecha_str:
                    try:
                        if isinstance(fecha_str, datetime):
                            fecha_procesada = fecha_str
                        elif isinstance(fecha_str, str):
                            # Intentar múltiples formatos de fecha
                            formatos_fecha = [
                                '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', 
                                '%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M'
                            ]
                            for formato in formatos_fecha:
                                try:
                                    fecha_procesada = datetime.strptime(fecha_str.strip(), formato)
                                    break
                                except ValueError:
                                    continue
                    except Exception as e:
                        print(f"⚠️ Error procesando fecha fila {row_num}: {str(e)}")
                
                # Usar fecha actual si no se pudo procesar
                if not fecha_procesada:
                    fecha_procesada = datetime.now()
                
                datos_excel.append({
                    'nombre': nombre,
                    'telefono': telefono,
                    'fecha_registro': fecha_procesada,
                    'fila': row_num
                })
        
        # ✅ DISTRIBUCIÓN EQUITATIVA MEJORADA
        if not datos_excel:
            return {
                "success": False,
                "message": "No hay datos válidos para procesar",
                "errores_detalle": errores[:10]
            }
        
        # Distribuir entre asesores activos
        distribucion = distribuir_equitativamente(datos_excel, asesores)
        
        # Crear reclutas en BD con manejo de errores robusto
        reclutas_creados = 0
        errores_bd = []
        
        for asesor in asesores:
            reclutas_asesor = distribucion.get(asesor.email, [])
            
            for datos_recluta in reclutas_asesor:
                try:
                    nuevo_recluta = Recluta(
                        nombre=datos_recluta['nombre'],
                        email=f"temp_{datos_recluta['telefono']}@temp.com",  # Email temporal
                        telefono=datos_recluta['telefono'],
                        estado='En proceso',
                        asesor_id=asesor.id,
                        fecha_registro=datos_recluta['fecha_registro']
                    )
                    nuevo_recluta.save()
                    reclutas_creados += 1
                    
                except Exception as e:
                    errores_bd.append({
                        "fila": datos_recluta['fila'], 
                        "error": f"Error BD: {str(e)}",
                        "datos": datos_recluta
                    })
        
        # Preparar reporte final detallado
        todos_errores = errores + errores_bd
        
        reporte_distribucion = {
            asesor.email: len(distribucion.get(asesor.email, []))
            for asesor in asesores
        }
        
        return {
            "success": True,
            "total_procesados": len(datos_excel) + len(errores),
            "exitosos": reclutas_creados,
            "errores": len(todos_errores),
            "distribucion": reporte_distribucion,
            "errores_detalle": todos_errores[:10],  # Primeros 10 errores para UI
            "resumen": {
                "filas_excel": row_num - 1,  # Total de filas procesadas
                "datos_validos": len(datos_excel),
                "datos_invalidos": len(errores),
                "errores_bd": len(errores_bd),
                "asesores_utilizados": len([a for a in asesores if a.email in distribucion])
            }
        }
        
    except Exception as e:
        return {
            "success": False,
            "message": f"Error crítico al procesar Excel: {str(e)}",
            "tipo_error": "error_sistema"
        }


def distribuir_equitativamente(datos_excel, asesores):
    """
    ✅ FUNCIÓN MEJORADA: Distribuye lista de datos equitativamente entre asesores.
    
    Args:
        datos_excel: Lista de datos de reclutas validados
        asesores: Lista de objetos Usuario asesor activos
        
    Returns:
        Dict con distribución por email de asesor
    """
    if not asesores or not datos_excel:
        return {}
    
    total_reclutas = len(datos_excel)
    num_asesores = len(asesores)
    
    # Cálculo de distribución equitativa
    reclutas_por_asesor = total_reclutas // num_asesores
    sobrantes = total_reclutas % num_asesores
    
    # Distribuir con algoritmo round-robin optimizado
    distribucion = {}
    indice_actual = 0
    
    for i, asesor in enumerate(asesores):
        # Asignar cantidad base + 1 extra para los primeros N asesores (sobrantes)
        cantidad_asignar = reclutas_por_asesor + (1 if i < sobrantes else 0)
        
        # Extraer slice correspondiente
        reclutas_asesor = datos_excel[indice_actual:indice_actual + cantidad_asignar]
        distribucion[asesor.email] = reclutas_asesor
        
        indice_actual += cantidad_asignar
        
        print(f"📊 Asesor {asesor.email}: {len(reclutas_asesor)} reclutas asignados")

    return distribucion